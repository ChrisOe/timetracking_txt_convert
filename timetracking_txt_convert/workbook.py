import openpyxl
from employee import Employee
from openpyxl.styles import Font, Alignment, Border, Side
# from openpyxl.utils import get_column_letter

OUTPUT_PATH = "output"


class Workbook:

    def __init__(self, month: str, year: str, department: str):
        self.workbook = openpyxl.Workbook()
        self.work_month = month
        self.work_year = year
        self.department = department
        print(f"Erstellung xlsx für {department} {self.work_month}-{self.work_year}.")
        self.sum_sheet = self.workbook.active
        self.sum_sheet.column_dimensions["A"].width = 16
        self.sum_sheet.title = f"{department}"
        self.sum_row = 0
        self.error_message = ""

    def add_employee_sheet(self, employee: Employee):
        print(f"    Hinzufügen Arbeitsblatt für {employee.full_name}.")
        self.sum_row += 1
        self.sum_sheet.cell(row=self.sum_row, column=1).value = employee.full_name
        self.sum_sheet.cell(row=self.sum_row, column=1).font = Font(bold=True)
        self.sum_row += 1
        self.sum_sheet.cell(row=self.sum_row, column=2).value = f"Personal-nr."
        self.sum_sheet.cell(row=self.sum_row, column=2).alignment = Alignment(horizontal="right")
        self.sum_sheet.cell(row=self.sum_row, column=3).value = int(employee.personal_no)
        self.sum_sheet.cell(row=self.sum_row, column=3).alignment = Alignment(horizontal="right")
        self.sum_row = self.add_employee_extra_days(employee, self.sum_row, 2, self.sum_sheet)
        sheet = self.workbook.create_sheet(employee.full_name)
        sheet.column_dimensions["A"].width = 10
        row = 1
        sheet.cell(row=row, column=1).value = employee.full_name
        sheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        sheet.cell(row=row, column=1).value = f"Personal-nr. {employee.personal_no}"
        sheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        row = self.add_employee_extra_days(employee, row, 3, sheet)
        for key, day in employee.days.items():
            row += 1
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            sheet.cell(row=row, column=1).value = key
            sheet.cell(row=row, column=1).font = Font(bold=True)
            sheet.cell(row=row, column=1).number_format = "DDDD dd.mm.yyyy"
            sheet.cell(row=row, column=1).alignment = Alignment(horizontal="left")
            row += 1
            # time-list header
            sheet.cell(row=row, column=1).value = "Art"
            sheet.cell(row=row, column=2).value = "Start"
            sheet.cell(row=row, column=3).value = "Ende"
            sheet.cell(row=row, column=4).value = "Dauer"
            for i in range(4):
                sheet.cell(row=row, column=i+1).font = Font(bold=True)
                sheet.cell(row=row, column=i+1).alignment = Alignment(horizontal="right")
                sheet.cell(row=row, column=i+1).border = Border(bottom=Side(border_style='medium'))
            # insert times
            for times in day["times"]:
                row += 1
                sheet.cell(row=row, column=1).value = times.time_type
                sheet.cell(row=row, column=1).alignment = Alignment(horizontal="right")
                sheet.cell(row=row, column=2).value = times.start
                sheet.cell(row=row, column=3).value = times.end
                sheet.cell(row=row, column=4).value = times.duration
                if not times.comment == "":
                    sheet.cell(row=row, column=5).value = times.comment
                for i in range(3):
                    sheet.cell(row=row, column=i+2).number_format = "hh:mm"
            for i in range(3):
                sheet.cell(row=row, column=i+1).border = Border(bottom=Side(border_style='thin'))
            row += 1
            sheet.cell(row=row, column=3).value = "gewertete Tagessumme:"
            sheet.cell(row=row, column=3).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=4).value = day["sum"]
            sheet.cell(row=row, column=4).number_format = "[h]:mm"
            sheet.cell(row=row, column=4).border = Border(top=Side(border_style='thin'),
                                                          bottom=Side(border_style='double'))
            row += 1

    def add_employee_extra_days(self, employee: Employee, row: int, col: int, sheet) -> int:
        extra_paid_days = 0
        if not employee.is_full_month:
            row += 1
            sheet.cell(row=row, column=col).value = f"MONAT {employee.report_month}-{employee.report_year} " \
                                                    f"NICHT VOLLSTÄNDIG ENTHALTEN!"
            sheet.cell(row=row, column=col).font = Font(color="FF0000", bold=True)
        if employee.holidays > 0:
            row += 1
            sheet.cell(row=row, column=col).value = "Feiertage:"
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=col + 1).value = employee.holidays
            extra_paid_days += employee.holidays
        if employee.days_paid_sick > 0:
            row += 1
            sheet.cell(row=row, column=col).value = "Krank mit LFZ:"
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=col + 1).value = employee.days_paid_sick
            extra_paid_days += employee.days_paid_sick
        if employee.days_unpaid_sick > 0:
            row += 1
            sheet.cell(row=row, column=col).value = "Krank ohne LFZ:"
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=col + 1).value = employee.days_unpaid_sick
        if employee.days_vacation > 0:
            row += 1
            sheet.cell(row=row, column=col).value = "Urlaub:"
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=col + 1).value = employee.days_vacation
            extra_paid_days += employee.days_vacation
        if employee.days_unpaid_vacation > 0:
            row += 1
            sheet.cell(row=row, column=col).value = "unbezahlter Urlaub:"
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=col + 1).value = employee.days_unpaid_vacation
        if extra_paid_days > 0 and employee.paid_per_hour:
            row += 1
            sheet.cell(row=row, column=col).value = "Zusätzlich bezahlte Tage:"
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=col + 1).value = extra_paid_days
        if employee.get_monthly_sum().seconds > 0:
            row += 1
            sheet.cell(row=row, column=col + 1).value = "hh:mm"
            sheet.cell(row=row, column=col + 2).value = "Dezimal"
            sheet.cell(row=row, column=col + 1).border = Border(bottom=Side(border_style="thin"))
            sheet.cell(row=row, column=col + 2).border = Border(bottom=Side(border_style="thin"))
            row += 1
            sheet.cell(row=row, column=col).value = "Monatsstunden:"
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=col + 1).value = employee.get_monthly_sum()
            sheet.cell(row=row, column=col + 1).number_format = "[h]:mm"
            month_sum_hours = employee.get_monthly_sum().seconds / 3600 + (employee.get_monthly_sum().days * 24)
            sheet.cell(row=row, column=col + 2).value = round(month_sum_hours, 2)
        month_mean = employee.three_month_mean.seconds + employee.three_month_mean.days * 86400
        if month_mean > 0:
            row += 1
            sheet.cell(row=row, column=col).value = "Dreimonatsmittel:"
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=col + 1).value = employee.three_month_mean
            sheet.cell(row=row, column=col + 1).number_format = "[h]:mm"
            sheet.cell(row=row, column=col + 2).value = round(month_mean / 3600, 2)
        if extra_paid_days > 0 and month_mean > 0:
            row += 1
            time_extra_days = employee.three_month_mean * extra_paid_days
            time_end_sum = time_extra_days + employee.get_monthly_sum()
            time_end_sum_decimal = time_end_sum.seconds / 3600 + time_end_sum.days * 24
            sheet.cell(row=row, column=col).value = "Zusätzliche Tage:"
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=col + 1).value = time_extra_days
            sheet.cell(row=row, column=col + 1).number_format = "[h]:mm"
            sheet.cell(row=row, column=col + 2).value = round(month_mean * extra_paid_days / 3600, 2)
            sheet.cell(row=row, column=col + 1).border = Border(bottom=Side(border_style="thin"))
            sheet.cell(row=row, column=col + 2).border = Border(bottom=Side(border_style="thin"))
            row += 1
            sheet.cell(row=row, column=col).value = "Summe:"
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=col + 1).value = time_end_sum
            sheet.cell(row=row, column=col + 1).number_format = "[h]:mm"
            sheet.cell(row=row, column=col + 2).value = round(time_end_sum_decimal, 2)
        row += 1
        return row

    def save_workbook(self) -> str:
        file_name = f"{self.work_year}-{int(self.work_month):02} Zeiten {self.department}.xlsx"
        try:
            self.workbook.save(f"{OUTPUT_PATH}/{file_name}")
            print(f"Speichern von Datei {file_name}.")
        except PermissionError:
            self.error_message += f"Speichern von Datei {file_name} nicht möglich.\n"
        return self.error_message
